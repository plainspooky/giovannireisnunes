from datetime import datetime
from random import randint
from tempfile import NamedTemporaryFile

from flask import Flask, Response
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

RAND = (1, 1000)
ROWS, COLS = 100, 100

DATE_FORMAT = "%Y%m%d_%H%M%S"

HTTP_MIMETYPE_XLSX = (
    "application/" "vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
)

app = Flask(__name__)


@app.route("/ws/", methods=["GET"])
def worksheet():

    wb = Workbook()
    ws = wb.active

    just_now = datetime.now()

    with NamedTemporaryFile() as xlsx:
        for row in range(1, ROWS + 1):
            for col in range(1, COLS + 1):
                ws["{}{}".format(get_column_letter(col), row)] = randint(*RAND)

        ws.title = just_now.strftime(DATE_FORMAT)
        wb.save(xlsx)

        xlsx.seek(0)

        return Response(xlsx.read(), mimetype=HTTP_MIMETYPE_XLSX)

#!/usr/bin/env python3
from csv import reader

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

STUDENTS = "students"

width = {}

wb = Workbook()
ws = wb.active

with open(STUDENTS + ".csv", "r") as f:
    for r, student in enumerate(reader(f), start=1):
        for c, data in enumerate(student, start=1):
            ws.cell(row=r, column=c, value=data)
            width[c] = max((width.get(c, 0), len(data)))

for c, size in width.items():
    ws.column_dimensions[get_column_letter(c)].width = 1 + size

ws.title = STUDENTS
wb.save(STUDENTS + ".xlsx")

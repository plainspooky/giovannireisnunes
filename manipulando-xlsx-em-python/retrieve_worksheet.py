#!/usr/bin/env python3
from colorama import Back, Fore, init
from openpyxl import load_workbook

ROWS, COLS = 10, 10
WORKBOOK, WORKSHEET = "aleatórios.xlsx", "Aleatório"

wb = load_workbook(WORKBOOK)
ws = wb[wb.sheetnames[0]]  # WORKSHEET]

for r in range(1, ROWS + 1):
    print(
        Fore.CYAN if r % 2 else Fore.MAGENTA,
        " | ".join(
            (
                "{:5d}".format(
                    ws.cell(row=r, column=c).value
                )
                for c in range(1, COLS + 1)
            )
        ),
        Fore.RESET,
        sep="",
    )
